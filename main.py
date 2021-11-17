import openpyxl
import xlrd
import xlwt
wb=openpyxl.load_workbook('list.xlsx')
names=wb.sheetnames
sheet = wb[names[0]]
maxrow = sheet.max_row
short_name=[]
full_name=[]
text=[]
leveal=['A','B','C']
index=2
for i in range(1,maxrow+1):
    if sheet.cell(row=i,column=2).value=='会议简称':
        index=(index+1)%3
    if sheet.cell(row=i,column=2).value!=None and sheet.cell(row=i,column=2).value!='会议简称':
        if sheet.cell(row=i,column=2).value=='ECML-PKDD':short_name.append({leveal[index]:'ECML PKDD'})
        else:short_name.append({leveal[index]:sheet.cell(row=i,column=2).value})
        full_name.append({leveal[index]:sheet.cell(row=i,column=3).value})
for item in full_name:
    for key,value in item.items():
        print(key+':'+value)
wb=openpyxl.load_workbook('journals.xlsx')
names=wb.worksheets
sheet=wb['Sheet1']
maxrow = sheet.max_row
workbook = xlwt.Workbook(encoding = 'utf-8')
worksheet = workbook.add_sheet('Sheet1')
for i in range(1,maxrow+1):

    flag=0
    if sheet.cell(row=i, column=5).value != None:
        for item in short_name:
            for key,avalue in item.items():
                if avalue in sheet.cell(row=i,column=5).value and flag!=1:
                    flag=1
                    worksheet.write(i, 0, label=key)

        for item in full_name:
            for key, avalue in item.items():
                if avalue in sheet.cell(row=i, column=5).value and flag!=1:
                    flag = 1
                    worksheet.write(i, 0, label=key)

    if flag==0:
        worksheet.write(i, 0, label=' ')
workbook.save('Excel_test.xlsx')
